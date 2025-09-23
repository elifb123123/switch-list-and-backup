import os
from openpyxl import Workbook, load_workbook

file_name = "data.xlsx"
  # You can change this

def addExcel(ad,marka,model,lokasyon, ip ,username, password):

    new_data = [ad, marka , model, lokasyon, ip ,username, password]
   
    if not os.path.exists(file_name):
        # If file does not exist, create a new one
        wb = Workbook()
        ws = wb.active
        ws.append(["Switch Name","Marka" ,"Model", "Lokasyon", "IP", "UserName", "Password" ])  # Headers
        ws.append(new_data)                             # First row of data
        wb.save(file_name)
        print("New file created and data added.")
    else:
        # If file exists, open and add data
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append(new_data)
        wb.save(file_name)
        print("Data appended to existing file.")


