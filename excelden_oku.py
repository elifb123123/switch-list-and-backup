from PyQt5.QtWidgets import QPushButton, QTableWidgetItem
from PyQt5.QtWidgets import qApp
from PyQt5.QtWidgets import QTableWidget
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter
from functools import partial
from paramikoo import connect_and_run
from PyQt5.QtWidgets import QMessageBox, QApplication
import paramiko
import socket
from PyQt5.QtCore import Qt



def load_excel_to_table(table_widget, file_name):

    if not os.path.exists(file_name):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Switch Adı", "Marka", "Model", "Lokasyon", "IP", "UserName", "Password"])# İstersen buraya varsayılan başlık ekleyebilirsin
            workbook.save(file_name)
        
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active

    row_count = sheet.max_row
    col_count = sheet.max_column

    table_widget.setRowCount(row_count - 1)
    table_widget.setColumnCount(col_count + 1)

    headers = [str(sheet.cell(row=1, column=col).value) for col in range(1, col_count )]
    headers.append("")
    headers.append("")
    table_widget.setHorizontalHeaderLabels(headers)

    sil_buttons = []
    backup_buttons= []

    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            value = sheet.cell(row=row, column=col).value
            item = QTableWidgetItem(str(value) if value is not None else "")
            table_widget.setItem(row - 2, col - 1, item)

        btn = QPushButton("Sil")
        sil_buttons.append(btn)
        table_widget.setCellWidget(row - 2, col_count-1, sil_buttons[row-2])
        sil_buttons[row-2].clicked.connect(partial(remove_row_and_update_excel, table_widget, row, workbook, sheet, file_name))

        btn2 = QPushButton("Backup")
        backup_buttons.append(btn2)
        table_widget.setCellWidget(row - 2, col_count, backup_buttons[row-2])
        backup_buttons[row-2].clicked.connect(partial(backup, table_widget, row, workbook, sheet, file_name))




def remove_row_and_update_excel(table_widget, row, workbook, sheet, file_name, parent=None):
    # Onay kutusu goster

    name = sheet.cell(row=row, column=1).value
    marka = sheet.cell(row=row, column=2).value
    model = sheet.cell(row=row, column=3).value
    lokasyon = sheet.cell(row=row, column=4).value
    host = sheet.cell(row=row, column=5).value

    

    msg = QMessageBox(table_widget.parent())
    msg.setIcon(QMessageBox.Warning)
    msg.setWindowTitle("Silme Onay\u0131")
    msg.setText("Switch ad:"+name+"\nMarka:"+marka+"\nModel:"+model+"\nLokasyon:"+lokasyon+"\nIP:"+host+"\n Bu switchi silmek istedi\u011finize emin misiniz?")
    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    msg.setDefaultButton(QMessageBox.No)

    yes_button = msg.button(QMessageBox.Yes)
    yes_button.setText("Evet")
    no_button = msg.button(QMessageBox.No)
    no_button.setText("Hay\u0131r")

    cevap = msg.exec_()

    if cevap == QMessageBox.Yes:
        # Excel'den sil
        sheet.delete_rows(row, 1)
        workbook.save(file_name)

        # Tabloyu güncelle
        load_excel_to_table(table_widget, file_name)



def backup(table_widget, row, workbook, sheet, file_name):
    # Implemented backup logic here
    host = sheet.cell(row=row, column=5).value
    username = sheet.cell(row=row, column=6).value
    password = sheet.cell(row=row, column=7).value
    try:
        connect_and_run(host, username, password, " show running-config")
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("   ")
        msg.setText(f"Backup ger\u00e7ekle\u015Fti.\n{host.replace('.', '_')}_output.txt olu\u015Fturuldu.")
        msg.setTextInteractionFlags(Qt.TextSelectableByMouse)
        msg.exec_()

    except paramiko.AuthenticationException:
        show_error_message(f"Authentication failed on {host}.")
    except paramiko.SSHException as e:
        show_error_message(f"SSH error on {host}: {e}")
    except socket.gaierror as e:
        show_error_message(f"Hostname \u00e7\u00F6z\u00FCmlenemedi: {e}")
    except Exception as e:
        show_error_message(f"Other error on {host}: {e}")

    
   
    load_excel_to_table(table_widget, file_name)





def show_error_message(text):
    app = QApplication.instance()
    if not app:
        app = QApplication([])
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setWindowTitle("Hata")
    msg.setText(text)
    msg.setTextInteractionFlags(Qt.TextSelectableByMouse)
    msg.exec_()





